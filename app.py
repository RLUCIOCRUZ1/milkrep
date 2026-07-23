import io
import logging
import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Font

from colunas import COLUNAS_EXTRAS_PEDIDO, alinhar_colunas_extras, mapear_colunas_clientes
from database import (
    carregar_vw_pedido_itens_filtrado,
    limpar_dados_automacao,
    montar_comissao_com_preposto,
    montar_pedidos_com_preposto,
    salvar_comissao,
    salvar_itens,
    salvar_pedidos,
)
from excel_colunas import (
    RENOME_COLUNAS_CANONICAS_SUPABASE,
    enriquecer_pedidos_colunas_excel,
    garantir_coluna_sku_por_letra,
    indices_excel_somente_pedidos,
)
from mailer import enviar_email_com_anexo, validar_config_smtp
from sheets import ler_dados, ler_lista_email
from utils import tratar_status

st.set_page_config(page_title="Milkyrep", layout="wide")

logger = logging.getLogger("milkyrep.app")
if not logger.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )

# Em produção (Render/Linux), caminhos absolutos do Windows quebram.
# Tenta usar um caminho configurado por variável de ambiente ou arquivos locais conhecidos.
_logo_env_path = os.getenv("APP_LOGO_PATH", "").strip().replace("\\", "/")
_logo_candidates = [
    _logo_env_path,
    str(Path(__file__).resolve().parent / _logo_env_path) if _logo_env_path else "",
    str(Path(__file__).with_name("logo.png")),
    str(Path(__file__).with_name("assets") / "logo.png"),
]
_logo_url = os.getenv("APP_LOGO_URL", "").strip()
_logo_source = _logo_url
if not _logo_source:
    for _logo_path in _logo_candidates:
        if _logo_path and Path(_logo_path).is_file():
            _logo_source = _logo_path
            break
if _logo_source:
    st.logo(_logo_source, size="large")

st.markdown(
    """
<style>
    .block-container {
        padding-top: 1rem;
        padding-left: 1.5rem;
        padding-right: 1.5rem;
    }
    div[data-testid="stButton"] > button,
    div[data-testid="stDownloadButton"] > button {
        border-radius: 14px !important;
        border: 1px solid #1e4f8d !important;
        background: linear-gradient(135deg, #0a3f7a 0%, #145ea8 100%) !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        box-shadow: 0 8px 20px rgba(10, 63, 122, 0.25) !important;
        transition: transform 0.18s ease, box-shadow 0.18s ease, filter 0.18s ease !important;
    }
    div[data-testid="stButton"] > button:hover,
    div[data-testid="stDownloadButton"] > button:hover {
        transform: translateY(-2px) scale(1.01) !important;
        box-shadow: 0 12px 24px rgba(10, 63, 122, 0.35) !important;
        filter: brightness(1.06) !important;
    }
    section[data-testid="stSidebar"] {
        border-right: 1px solid rgba(20, 94, 168, 0.12);
    }
    section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] img {
        margin-bottom: 12px;
    }
</style>
""",
    unsafe_allow_html=True,
)
COLUNAS_OBRIGATORIAS = ("customer", "store", "customer_name", "order_no", "style", "rsn")
# Ordem das colunas no Excel enviado por e-mail (só as que existirem na view).
COLUNAS_EXCEL_CARTEIRA = (
    "store",
    "nome_fantasia",
    "cnpj",
    "customer_name",
    "style",
    "descricao_modelo",
    "color",
    "genero",
    "preco_liquido",
    "total",
    "data_faturamento",
    "tamanho",
    "quantidade",
    "preposto",
)
ROTULOS_EXCEL_CARTEIRA = {
    "style": "referencia",
    "color": "cor",
}
# Valores reais no banco: MENS / WOMENS / BOYS / GIRLS
TRADUCAO_GENERO = {
    "mens": "Masculino",
    "womens": "Feminino",
    "boys": "Infantil masculino",
    "girls": "Infantil feminino",
}
ORDEM_GENERO_RESUMO = (
    "Masculino",
    "Feminino",
    "Infantil masculino",
    "Infantil feminino",
)
ASSINATURA_EMAIL = (
    "Atenciosamente,\n\n"
    "Maísa Gomes\n"
    "MILKY REPRESENTAÇÕES COMERCIAIS LTDA\n"
    "(62) 3271-1026 / whatsapp 99275-3077"
)


def _to_float_series(serie: pd.Series) -> pd.Series:
    s = serie.astype(str).str.strip()
    s = s.replace({"": None, "None": None, "nan": None})
    s = s.str.replace(r"[^\d,.\-]", "", regex=True)
    mask_both = s.str.contains(",", na=False) & s.str.contains(r"\.", na=False)
    s.loc[mask_both] = (
        s.loc[mask_both]
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    mask_comma = s.str.contains(",", na=False) & ~s.str.contains(r"\.", na=False)
    s.loc[mask_comma] = s.loc[mask_comma].str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0)


def _format_brl(valor: float) -> str:
    txt = f"{float(valor):,.2f}"
    txt = txt.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {txt}"


def _format_pecas(valor: float) -> str:
    v = float(valor)
    if abs(v - round(v)) < 1e-9:
        return f"{int(round(v))} peças"
    txt = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{txt} peças"


def _normalizar_texto(valor: str) -> str:
    txt = unicodedata.normalize("NFKD", str(valor))
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return txt.lower().strip()


def _status_para_resumo(status_original: str) -> tuple[str, int, int]:
    meses = {
        "janeiro": 1,
        "jan": 1,
        "fevereiro": 2,
        "fev": 2,
        "marco": 3,
        "mar": 3,
        "abril": 4,
        "abr": 4,
        "maio": 5,
        "mai": 5,
        "junho": 6,
        "jun": 6,
        "julho": 7,
        "jul": 7,
        "agosto": 8,
        "ago": 8,
        "setembro": 9,
        "set": 9,
        "outubro": 10,
        "out": 10,
        "novembro": 11,
        "nov": 11,
        "dezembro": 12,
        "dez": 12,
    }
    nomes_meses = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }
    texto = str(status_original or "").strip()
    txt_norm = _normalizar_texto(texto)
    if txt_norm.startswith("liberacao"):
        mes_ordem = 99
        for token in txt_norm.split():
            if token in meses:
                mes_ordem = meses[token]
        if mes_ordem != 99:
            return f"Liberação {nomes_meses[mes_ordem]}", 0, mes_ordem
        return "Liberação (Sem mês)", 0, 99
    return texto or "Sem informação", 1, 999


def _traduzir_genero(valor) -> str:
    bruto = str(valor).strip() if valor is not None else ""
    if not bruto or bruto.lower() in {"nan", "none", "null"}:
        return "Sem informação"
    chave = _normalizar_texto(bruto)
    return TRADUCAO_GENERO.get(chave, bruto)


def _aplicar_metricas_qtd_preco(base: pd.DataFrame) -> pd.DataFrame:
    """Prepara quantidade e preço líquido para o resumo (valor = qtd × média do preço)."""
    out = base.copy()
    col_qtd = _resolver_coluna(out, ("quantidade", "qtd", "quantity"))
    col_preco = _resolver_coluna(out, ("preco_liquido", "preço_liquido", "price_liquido"))
    if col_qtd:
        out["__qtd_num"] = _to_float_series(out[col_qtd])
    else:
        out["__qtd_num"] = 0.0
    if col_preco:
        out["__preco_num"] = _to_float_series(out[col_preco])
    else:
        out["__preco_num"] = 0.0
    out["__valor_linha"] = out["__qtd_num"] * out["__preco_num"]
    return out


def _agregar_qtd_custo_valor(grupo: pd.DataFrame) -> pd.Series:
    """quantidade = soma; custo = média ponderada do preço líquido; valor = qtd × custo."""
    qtd = float(grupo["__qtd_num"].sum())
    valor_linhas = float(grupo["__valor_linha"].sum())
    if qtd > 0:
        custo = valor_linhas / qtd
    elif len(grupo):
        custo = float(grupo["__preco_num"].mean())
    else:
        custo = 0.0
    return pd.Series({"quantidade": qtd, "custo": custo, "valor": qtd * custo})


def _preparar_df_excel_carteira(df: pd.DataFrame) -> pd.DataFrame:
    """Seleciona e ordena colunas do anexo; traduz gênero; renomeia style/color."""
    base = df.copy()
    colunas_presentes: list[str] = []
    for nome in COLUNAS_EXCEL_CARTEIRA:
        col = _resolver_coluna(base, (nome,))
        if col is not None and col not in colunas_presentes:
            colunas_presentes.append(col)

    if not colunas_presentes:
        return base

    saida = base[colunas_presentes].copy()
    col_genero = _resolver_coluna(saida, ("genero",))
    if col_genero is not None:
        saida[col_genero] = saida[col_genero].map(_traduzir_genero)

    rename: dict[str, str] = {}
    for origem, rotulo in ROTULOS_EXCEL_CARTEIRA.items():
        col = _resolver_coluna(saida, (origem,))
        if col is not None:
            rename[col] = rotulo
    if rename:
        saida = saida.rename(columns=rename)
    return saida


def _montar_resumo_excel(df: pd.DataFrame) -> pd.DataFrame:
    base = _aplicar_metricas_qtd_preco(df)
    col_status = _resolver_coluna(base, ("status_pedido", "status"))
    col_modelo = _resolver_coluna(base, ("descricao_modelo",))

    if col_status is None and "rsn" in base.columns:
        base["status_pedido"] = base.apply(
            lambda row: tratar_status(row.get("rsn"), row.get("pick_date")), axis=1
        )
        col_status = "status_pedido"
    if col_status is None:
        base["status_pedido"] = "Sem informação"
        col_status = "status_pedido"

    if col_modelo is None:
        base["descricao_modelo"] = "Sem informação"
        col_modelo = "descricao_modelo"

    status_meta = base[col_status].map(_status_para_resumo)
    base["status_resumo"] = status_meta.map(lambda x: x[0])
    base["grupo_tipo"] = status_meta.map(lambda x: x[1])
    base["ordem_mes"] = status_meta.map(lambda x: x[2])

    status_ordenado = (
        base.groupby("status_resumo", dropna=False, group_keys=False)
        .apply(_agregar_qtd_custo_valor, include_groups=False)
        .reset_index()
    )
    meta_status = (
        base.groupby("status_resumo", dropna=False, as_index=False)[["grupo_tipo", "ordem_mes"]]
        .min()
    )
    status_ordenado = status_ordenado.merge(meta_status, on="status_resumo", how="left")
    status_ordenado = status_ordenado.sort_values(
        ["grupo_tipo", "ordem_mes", "status_resumo"], ascending=True
    )

    detalhes_modelo = (
        base.groupby(["status_resumo", col_modelo], dropna=False, group_keys=False)
        .apply(_agregar_qtd_custo_valor, include_groups=False)
        .reset_index()
        .rename(columns={col_modelo: "descricao_modelo"})
        .sort_values(
            ["status_resumo", "valor", "quantidade"], ascending=[True, False, False]
        )
    )

    linhas_resumo: list[dict[str, str]] = []
    for _, row_status in status_ordenado.iterrows():
        status_nome = row_status["status_resumo"]
        linhas_resumo.append(
            {
                "status_pedido": status_nome,
                "quantidade": _format_pecas(row_status["quantidade"]),
                "custo": _format_brl(row_status["custo"]),
                "valor": _format_brl(row_status["valor"]),
            }
        )
        detalhes_status = detalhes_modelo[detalhes_modelo["status_resumo"] == status_nome]
        for row_modelo in detalhes_status.itertuples(index=False):
            nome_modelo = str(row_modelo.descricao_modelo).strip() or "Sem informação"
            linhas_resumo.append(
                {
                    "status_pedido": f"    {nome_modelo}",
                    "quantidade": _format_pecas(row_modelo.quantidade),
                    "custo": _format_brl(row_modelo.custo),
                    "valor": _format_brl(row_modelo.valor),
                }
            )
        linhas_resumo.append(
            {"status_pedido": "", "quantidade": "", "custo": "", "valor": ""}
        )

    if linhas_resumo and not linhas_resumo[-1]["status_pedido"]:
        linhas_resumo.pop()

    return pd.DataFrame(
        linhas_resumo, columns=["status_pedido", "quantidade", "custo", "valor"]
    )


def _montar_resumo_genero_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Resumo por gênero, com detalhe por mês de liberação (como descrição no status)."""
    base = _aplicar_metricas_qtd_preco(df)
    col_genero = _resolver_coluna(base, ("genero",))
    col_status = _resolver_coluna(base, ("status_pedido", "status"))

    if col_genero is None:
        base["genero"] = "Sem informação"
        col_genero = "genero"

    if col_status is None and "rsn" in base.columns:
        base["status_pedido"] = base.apply(
            lambda row: tratar_status(row.get("rsn"), row.get("pick_date")), axis=1
        )
        col_status = "status_pedido"
    if col_status is None:
        base["status_pedido"] = "Sem informação"
        col_status = "status_pedido"

    base["genero_resumo"] = base[col_genero].map(_traduzir_genero)
    status_meta = base[col_status].map(_status_para_resumo)
    base["status_resumo"] = status_meta.map(lambda x: x[0])
    base["grupo_tipo"] = status_meta.map(lambda x: x[1])
    base["ordem_mes"] = status_meta.map(lambda x: x[2])

    ordem_genero = {nome: i for i, nome in enumerate(ORDEM_GENERO_RESUMO)}
    genero_ordenado = (
        base.groupby("genero_resumo", dropna=False, group_keys=False)
        .apply(_agregar_qtd_custo_valor, include_groups=False)
        .reset_index()
    )
    genero_ordenado["__ordem"] = genero_ordenado["genero_resumo"].map(
        lambda g: ordem_genero.get(str(g), 100)
    )
    genero_ordenado = genero_ordenado.sort_values(
        ["__ordem", "genero_resumo"], ascending=True
    )

    detalhes_liberacao = (
        base.groupby(
            ["genero_resumo", "status_resumo", "grupo_tipo", "ordem_mes"],
            dropna=False,
            group_keys=False,
        )
        .apply(_agregar_qtd_custo_valor, include_groups=False)
        .reset_index()
        .sort_values(
            ["genero_resumo", "grupo_tipo", "ordem_mes", "status_resumo"],
            ascending=True,
        )
    )

    linhas: list[dict[str, str]] = []
    for row_genero in genero_ordenado.itertuples(index=False):
        genero_nome = str(row_genero.genero_resumo).strip() or "Sem informação"
        linhas.append(
            {
                "genero": genero_nome,
                "quantidade": _format_pecas(row_genero.quantidade),
                "custo": _format_brl(row_genero.custo),
                "valor": _format_brl(row_genero.valor),
            }
        )
        detalhes = detalhes_liberacao[detalhes_liberacao["genero_resumo"] == genero_nome]
        for row_lib in detalhes.itertuples(index=False):
            nome_lib = str(row_lib.status_resumo).strip() or "Sem informação"
            linhas.append(
                {
                    "genero": f"    {nome_lib}",
                    "quantidade": _format_pecas(row_lib.quantidade),
                    "custo": _format_brl(row_lib.custo),
                    "valor": _format_brl(row_lib.valor),
                }
            )
        linhas.append({"genero": "", "quantidade": "", "custo": "", "valor": ""})

    if linhas and not linhas[-1]["genero"]:
        linhas.pop()

    return pd.DataFrame(linhas, columns=["genero", "quantidade", "custo", "valor"])


def _excel_bytes(df: pd.DataFrame, sheet_name: str) -> bytes:
    # Resumos usam o DF completo (precisa de status_pedido etc.).
    resumo_df = _montar_resumo_excel(df)
    resumo_genero_df = _montar_resumo_genero_excel(df)
    df_anexo = _preparar_df_excel_carteira(df)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_anexo.to_excel(writer, sheet_name=sheet_name, index=False)
        resumo_df.to_excel(writer, sheet_name="resumo", index=False, startrow=0, startcol=0)
        # Resumo de gênero ao lado (coluna F = índice 5), com 1 coluna em branco.
        resumo_genero_df.to_excel(
            writer, sheet_name="resumo", index=False, startrow=0, startcol=5
        )
        ws_resumo = writer.book["resumo"]
        fonte_negrito = Font(bold=True)
        # Totais (status à esquerda, gênero à direita): sem indentação = negrito.
        for row_idx in range(2, ws_resumo.max_row + 1):
            for col_base in (1, 6):
                valor = ws_resumo.cell(row=row_idx, column=col_base).value
                if (
                    isinstance(valor, str)
                    and valor.strip()
                    and not valor.startswith("    ")
                ):
                    for col_idx in range(col_base, col_base + 4):
                        ws_resumo.cell(row=row_idx, column=col_idx).font = fonte_negrito
        for col_idx in (6, 7, 8, 9):
            ws_resumo.cell(row=1, column=col_idx).font = fonte_negrito
    buf.seek(0)
    return buf.read()


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower()).strip("_")


def _resolver_coluna(df: pd.DataFrame, aliases: tuple[str, ...]) -> str | None:
    mapa = {str(c): _slug(c) for c in df.columns}
    por_slug: dict[str, str] = {}
    for col, n in mapa.items():
        if n not in por_slug:
            por_slug[n] = col
    for alias in aliases:
        col = por_slug.get(_slug(alias))
        if col is not None:
            return col
    return None


def _key_customer(v) -> str | None:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"\s+", "", s)
    s_num = s.replace(",", ".")
    if re.fullmatch(r"\d+(\.\d+)?", s_num):
        try:
            n = float(s_num)
            if n.is_integer():
                return str(int(n))
        except ValueError:
            pass
    return s.upper()


def _key_store(v) -> str | None:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"\s+", "", s)
    s_num = s.replace(",", ".")
    if re.fullmatch(r"\d+(\.\d+)?", s_num):
        try:
            n = float(s_num)
            if n.is_integer():
                return str(int(n))
        except ValueError:
            pass
    return s.upper()


def _destinatarios(texto: str) -> list[str]:
    if not texto:
        return []
    parts = re.split(r"[;,]", str(texto))
    return [p.strip() for p in parts if p.strip() and "@" in p]


def _saudacao_horario() -> str:
    h = datetime.now().hour
    if h < 12:
        return "Bom dia"
    if h < 18:
        return "Boa tarde"
    return "Boa noite"


def _referencia_cliente(df: pd.DataFrame, customer_chave: str) -> str:
    for coluna_nome in ("nome_fantasia", "customer_name"):
        if coluna_nome not in df.columns:
            continue
        nomes_validos = df[coluna_nome].dropna().map(lambda v: str(v).strip())
        nomes_validos = nomes_validos[
            nomes_validos.map(lambda v: bool(v) and v.lower() not in {"nan", "none"})
        ]
        if not nomes_validos.empty:
            return nomes_validos.iloc[0]
    return f"Customer {customer_chave}"


def _montar_corpo_email(referencia_cliente: str) -> str:
    return (
        f"{_saudacao_horario()},\n\n"
        "Segue em anexo sua carteira Skechers.\n"
        f"Cliente: {referencia_cliente}.\n\n"
        "Qualquer dúvida, estamos à disposição.\n\n"
        f"{ASSINATURA_EMAIL}"
    )


def _filtrar_df_por_customer_store(
    df_vw: pd.DataFrame, customer_chave: str, store_chave: str | None
) -> pd.DataFrame:
    store_chave = _key_store(store_chave)
    if store_chave:
        df_match = df_vw[
            (df_vw["_key_customer"] == customer_chave) & (df_vw["_key_store"] == store_chave)
        ]
        if not df_match.empty:
            return df_match
        # Fallback: compara o valor bruto de `store` como texto para cobrir diferenças de formatação.
        store_bruto = df_vw["store"].astype(str).str.strip().str.replace(r"\s+", "", regex=True).str.upper()
        return df_vw[(df_vw["_key_customer"] == customer_chave) & (store_bruto == str(store_chave).upper())]
    return df_vw[df_vw["_key_customer"] == customer_chave]


def _enviar_carteiras_email() -> None:
    inicio_exec = time.perf_counter()
    logger.info("Iniciando envio em lote de carteiras por e-mail.")
    ok_cfg, msg_cfg = validar_config_smtp()
    if not ok_cfg:
        logger.error("Config SMTP inválida para envio em lote: %s", msg_cfg)
        st.error(msg_cfg)
        st.info(
            "Para responderem ao e-mail, configure `SMTP_FROM` e `SMTP_REPLY_TO` no .env "
            "com uma caixa real monitorada."
        )
        return

    barra = st.progress(0)
    status = st.empty()

    def avancar(p: int, m: str) -> None:
        barra.progress(max(0, min(100, p)))
        status.caption(f"{p}% - {m}")

    avancar(5, "Lendo lista_email...")
    df_lista, _ = ler_lista_email()
    if df_lista.empty:
        logger.warning("Envio em lote abortado: aba lista_email vazia.")
        st.error("A aba `lista_email` esta vazia.")
        return

    col_customer = _resolver_coluna(df_lista, ("customer", "codigo_cliente", "cliente"))
    col_store = _resolver_coluna(df_lista, ("store", "loja", "codigo_loja", "cod_loja"))
    col_email = _resolver_coluna(df_lista, ("email", "e_mail", "mail"))
    if not col_customer or not col_email:
        logger.error("Envio em lote abortado: colunas customer/email não encontradas em lista_email.")
        st.error("Nao encontrei as colunas `customer` e `email` na aba `lista_email`.")
        return

    base_envio = pd.DataFrame(
        {
            "customer": df_lista[col_customer].map(_key_customer),
            "store": (
                df_lista[col_store].map(_key_store)
                if col_store
                else pd.Series([None] * len(df_lista), index=df_lista.index)
            ),
            "destinatarios": df_lista[col_email].map(_destinatarios),
        }
    ).dropna(subset=["customer"])
    base_envio = base_envio[base_envio["destinatarios"].map(bool)]
    base_envio = (
        base_envio.groupby(["customer", "store"], dropna=False, as_index=False)["destinatarios"]
        .agg(
            lambda listas: sorted(
                {
                    email.strip().lower()
                    for lista in listas
                    for email in lista
                    if isinstance(email, str) and email.strip()
                }
            )
        )
        .reset_index(drop=True)
    )
    if base_envio.empty:
        logger.warning("Envio em lote abortado: nenhum destinatário válido na lista_email.")
        st.error("Nenhum destinatario valido encontrado na aba `lista_email`.")
        return
    logger.info("Envio em lote iniciado com %s combinacoes customer/store elegíveis.", len(base_envio))

    avancar(20, "Carregando dados por cliente...")
    hoje = datetime.now().strftime("%d/%m/%Y")
    assunto = f"Carteira Skechers - {hoje}"
    enviados = 0
    sem_dados = 0
    erros: list[str] = []
    total = len(base_envio)

    for i, row in enumerate(base_envio.itertuples(index=False), start=1):
        customer = row.customer
        store = _key_store(row.store)
        destinatarios = row.destinatarios
        ini_customer = time.perf_counter()
        progresso = 20 + int((i / max(total, 1)) * 75)
        alvo = f"{customer}+{store}" if store else customer
        avancar(progresso, f"Enviando {i}/{total} para {alvo}...")

        try:
            df_vw, _ = carregar_vw_pedido_itens_filtrado(customer, store)
        except Exception as e:
            logger.exception("Falha ao carregar view para alvo %s: %s", alvo, e)
            erros.append(f"{alvo}: {e}")
            continue
        if df_vw.empty or "customer" not in df_vw.columns:
            sem_dados += 1
            logger.info("Alvo %s sem dados na view. Pulando.", alvo)
            continue
        df_vw = df_vw.copy()
        df_vw["_key_customer"] = df_vw["customer"].map(_key_customer)
        df_vw["_key_store"] = df_vw["store"].map(_key_store)
        df_cliente = _filtrar_df_por_customer_store(df_vw, customer, store).drop(
            columns=["_key_customer", "_key_store"]
        )
        if df_cliente.empty:
            sem_dados += 1
            logger.info("Alvo %s sem dados na view. Pulando.", alvo)
            continue

        anexo = _excel_bytes(df_cliente, "vw_pedidos_itens")
        sufixo_store = f"_{store}" if store else ""
        nome_anexo = f"carteira_skechers_{customer}{sufixo_store}_{datetime.now():%Y%m%d}.xlsx"
        referencia_cliente = _referencia_cliente(df_cliente, customer)
        corpo = _montar_corpo_email(referencia_cliente)
        try:
            logger.info(
                "Enviando alvo %s para %s destinatários (%s linhas).",
                alvo,
                len(destinatarios),
                len(df_cliente),
            )
            enviar_email_com_anexo(
                destinatarios=destinatarios,
                assunto=assunto,
                corpo_texto=corpo,
                anexo_bytes=anexo,
                anexo_nome=nome_anexo,
            )
            enviados += 1
            logger.info(
                "Envio alvo %s concluído em %.2fs.",
                alvo,
                time.perf_counter() - ini_customer,
            )
        except Exception as e:
            logger.exception("Falha no envio para alvo %s: %s", alvo, e)
            erros.append(f"{alvo}: {e}")

    avancar(100, "Processo de envio concluido.")
    st.success(
        f"Envio finalizado. Enviados: {enviados} | Sem dados: {sem_dados} | Erros: {len(erros)}"
    )
    if erros:
        st.warning("Falhas:\n- " + "\n- ".join(erros[:10]))
    logger.info(
        "Envio em lote finalizado em %.2fs | enviados=%s | sem_dados=%s | erros=%s",
        time.perf_counter() - inicio_exec,
        enviados,
        sem_dados,
        len(erros),
    )


def _enviar_carteira_por_customer(customer_informado: str, store_informado: str = "") -> None:
    inicio_exec = time.perf_counter()
    logger.info(
        "Iniciando envio por customer/store: customer=%s | store=%s",
        customer_informado,
        store_informado,
    )
    ok_cfg, msg_cfg = validar_config_smtp()
    if not ok_cfg:
        logger.error("Config SMTP inválida para envio por customer: %s", msg_cfg)
        st.error(msg_cfg)
        return

    customer_chave = _key_customer(customer_informado)
    store_chave = _key_store(store_informado)
    if not customer_chave:
        logger.warning("Envio por customer abortado: customer inválido.")
        st.error("Informe um customer valido para envio.")
        return

    barra = st.progress(0)
    status = st.empty()

    def avancar(p: int, m: str) -> None:
        barra.progress(max(0, min(100, p)))
        status.caption(f"{p}% - {m}")

    avancar(10, "Lendo lista_email...")
    df_lista, _ = ler_lista_email()
    if df_lista.empty:
        logger.warning("Envio por customer abortado: lista_email vazia.")
        st.error("A aba `lista_email` esta vazia.")
        return

    col_customer = _resolver_coluna(df_lista, ("customer", "codigo_cliente", "cliente"))
    col_store = _resolver_coluna(df_lista, ("store", "loja", "codigo_loja", "cod_loja"))
    col_email = _resolver_coluna(df_lista, ("email", "e_mail", "mail"))
    if not col_customer or not col_email:
        logger.error("Envio por customer abortado: colunas customer/email ausentes.")
        st.error("Nao encontrei as colunas `customer` e `email` na aba `lista_email`.")
        return

    base_envio = pd.DataFrame(
        {
            "customer": df_lista[col_customer].map(_key_customer),
            "store": (
                df_lista[col_store].map(_key_store)
                if col_store
                else pd.Series([None] * len(df_lista), index=df_lista.index)
            ),
            "destinatarios": df_lista[col_email].map(_destinatarios),
        }
    ).dropna(subset=["customer"])
    base_envio = base_envio[base_envio["destinatarios"].map(bool)]
    if store_chave:
        base_customer = base_envio[
            (base_envio["customer"] == customer_chave) & (base_envio["store"] == store_chave)
        ]
        rotulo_alvo = f"{customer_chave}+{store_chave}"
        msg_destinatario = (
            f"Nao encontrei destinatarios para `{rotulo_alvo}` na aba `lista_email`."
        )
    else:
        base_customer = base_envio[
            (base_envio["customer"] == customer_chave) & (base_envio["store"].isna())
        ]
        rotulo_alvo = customer_chave
        msg_destinatario = (
            f"Nao encontrei destinatarios para o customer `{customer_chave}` "
            "com store em branco na aba `lista_email`."
        )
    if base_customer.empty:
        logger.warning("Envio por customer/store abortado: %s sem destinatários.", rotulo_alvo)
        st.error(msg_destinatario)
        return

    destinatarios = sorted(
        {d for lista in base_customer["destinatarios"] for d in lista if d.strip()}
    )
    if not destinatarios:
        logger.warning("Envio por customer/store abortado: %s sem e-mails válidos.", rotulo_alvo)
        st.error(
            f"O alvo `{rotulo_alvo}` foi encontrado, mas sem e-mails validos para envio."
        )
        return

    avancar(35, "Carregando dados da carteira...")
    try:
        df_vw, nome_vw = carregar_vw_pedido_itens_filtrado(customer_chave, store_chave)
    except Exception as e:
        logger.exception("Falha ao carregar view para envio: %s", e)
        st.error(f"Nao foi possivel carregar a view de pedidos: {e}")
        return
    if df_vw.empty:
        logger.warning("Envio por customer abortado: view vazia para o alvo.")
        st.error("Nao ha dados na carteira para o customer/store informado.")
        return
    if "customer" not in df_vw.columns:
        logger.error("Envio por customer/store abortado: coluna customer ausente na view.")
        st.error("A view nao possui a coluna `customer` para filtrar o envio.")
        return
    if "store" not in df_vw.columns:
        logger.error("Envio por customer/store abortado: coluna store ausente na view.")
        st.error("A view nao possui a coluna `store` para filtrar o envio.")
        return

    df_vw = df_vw.copy()
    df_vw["_key_customer"] = df_vw["customer"].map(_key_customer)
    df_vw["_key_store"] = df_vw["store"].map(_key_store)
    df_customer = _filtrar_df_por_customer_store(df_vw, customer_chave, store_chave).drop(
        columns=["_key_customer", "_key_store"]
    )
    if df_customer.empty:
        logger.warning("Envio por customer/store abortado: %s sem linhas na view.", rotulo_alvo)
        stores_disponiveis = sorted(
            {
                str(v).strip()
                for v in df_vw.loc[df_vw["_key_customer"] == customer_chave, "store"].dropna().tolist()
                if str(v).strip()
            }
        )
        detalhe_store = ""
        if store_chave:
            detalhe_store = (
                f" Stores disponiveis para customer `{customer_chave}`: "
                + (", ".join(stores_disponiveis[:20]) if stores_disponiveis else "nenhum")
                + "."
            )
        st.error(
            f"Nao encontrei linhas na `vw_pedidos_itens` para `{rotulo_alvo}`.{detalhe_store}"
        )
        return

    avancar(60, "Gerando anexo Excel filtrado por customer/store...")
    anexo = _excel_bytes(df_customer, "vw_pedidos_itens")

    hoje = datetime.now().strftime("%d/%m/%Y")
    assunto = f"Carteira Skechers - {hoje}"
    referencia_cliente = _referencia_cliente(df_customer, customer_chave)
    corpo = _montar_corpo_email(referencia_cliente)

    avancar(85, f"Enviando carteira para {rotulo_alvo}...")
    try:
        logger.info(
            "Enviando alvo %s para %s destinatários (%s linhas).",
            rotulo_alvo,
            len(destinatarios),
            len(df_customer),
        )
        enviar_email_com_anexo(
            destinatarios=destinatarios,
            assunto=assunto,
            corpo_texto=corpo,
            anexo_bytes=anexo,
            anexo_nome=(
                f"carteira_skechers_{customer_chave}"
                f"{'_' + store_chave if store_chave else ''}_{datetime.now():%Y%m%d_%H%M}.xlsx"
            ),
        )
    except Exception as e:
        logger.exception("Falha no envio por alvo %s: %s", rotulo_alvo, e)
        st.error(f"Falha no envio da carteira para `{rotulo_alvo}`: {e}")
        return

    avancar(100, "Envio concluido.")
    st.success(
        f"Carteira enviada com sucesso para `{rotulo_alvo}` "
        f"({len(destinatarios)} destinatario(s), {len(df_customer)} linha(s))."
    )
    logger.info(
        "Envio por customer/store finalizado em %.2fs para %s.",
        time.perf_counter() - inicio_exec,
        rotulo_alvo,
    )


def _executar_automacao_completa() -> None:
    barra = st.progress(0)
    status = st.empty()

    def avancar(percentual: int, msg: str) -> None:
        barra.progress(max(0, min(100, percentual)))
        status.caption(f"{percentual}% - {msg}")

    avancar(5, "Iniciando automacao...")

    (
        df_clientes,
        df_comissao,
        df_vendedores,
        _aba_clientes,
        _aba_comissao,
        _aba_vendedores,
    ) = ler_dados()
    avancar(20, "Planilhas carregadas.")

    if df_clientes.empty:
        st.error("A aba dados_clientes esta vazia ou sem linhas de dados.")
        return

    headers_orig = list(df_clientes.columns)
    df_clientes = mapear_colunas_clientes(df_clientes)
    alinhar_colunas_extras(df_clientes)
    df_clientes = garantir_coluna_sku_por_letra(df_clientes)
    avancar(35, "Colunas mapeadas e normalizadas.")

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in df_clientes.columns]
    if faltando:
        st.error(
            "Nao foi possivel identificar estas colunas na planilha: "
            + ", ".join(faltando)
        )
        return

    df_clientes["status_pedido"] = df_clientes["rsn"].apply(tratar_status)
    colunas_base = [
        "customer",
        "store",
        "customer_name",
        "order_no",
        "style",
        "rsn",
        "status_pedido",
    ]
    extras = [c for c in COLUNAS_EXTRAS_PEDIDO if c in df_clientes.columns]
    df_pedidos = df_clientes[colunas_base + extras]
    df_pedidos = df_pedidos.rename(columns=RENOME_COLUNAS_CANONICAS_SUPABASE)
    df_pedidos = enriquecer_pedidos_colunas_excel(
        df_clientes, df_pedidos, headers_orig, indices_excel_somente_pedidos()
    )
    df_pedidos["status_pedido"] = df_pedidos.apply(
        lambda row: tratar_status(row.get("rsn"), row.get("pick_date")),
        axis=1,
    )
    df_pedidos = montar_pedidos_com_preposto(df_pedidos, df_vendedores)
    avancar(55, "Dados de pedidos preparados.")

    limpar_dados_automacao()
    avancar(70, "Tabelas limpas.")
    salvar_pedidos(df_pedidos)
    avancar(82, "Pedidos gravados.")
    salvar_itens(df_clientes)
    avancar(90, "Itens gravados.")

    df_comissao_final = montar_comissao_com_preposto(df_comissao, df_vendedores)
    tabela_comissao, qtd_comissao = salvar_comissao(df_comissao_final)
    avancar(100, "Comissionamento atualizado.")

    st.success(
        "Automacao concluida com sucesso. "
        f"Comissao: {qtd_comissao} linhas em `{tabela_comissao}`."
    )


st.title("Milkyrep")
st.caption("Central de automacao e navegacao do projeto.")

if st.button("🚀 Rodar automacao completa (Pedidos + Comissao)", use_container_width=True):
    _executar_automacao_completa()

if st.button("📧 Enviar carteira por e-mail (lista_email)", use_container_width=True):
    _enviar_carteiras_email()

st.subheader("Envio por Customer")
c1_envio, c2_envio = st.columns(2)
with c1_envio:
    customer_envio = st.text_input(
        "Customer para envio",
        value="",
        help="Informe o customer para buscar os e-mails na aba `lista_email`.",
    )
with c2_envio:
    store_envio = st.text_input(
        "Store para envio",
        value="",
        help="Opcional. Se informado, envia por customer+store. Se vazio, usa customer com store em branco.",
    )
if st.button("🧪 Enviar carteira por customer", use_container_width=True):
    _enviar_carteira_por_customer(customer_envio, store_envio)

st.subheader("Navegação")
c1, c2 = st.columns(2)
with c1:
    if st.button("📦 Ir para Pedidos", use_container_width=True):
        st.switch_page("pages/1_Pedidos.py")
with c2:
    if st.button("💰 Ir para Comissionamento", use_container_width=True):
        st.switch_page("pages/2_Comissionamento.py")
